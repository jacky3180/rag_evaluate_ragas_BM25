"""
标准数据集构建模块
使用Ragas API构建标准数据集
"""

import os
import pandas as pd
from pathlib import Path
from typing import List, Dict, Any, Optional
from config import debug_print, verbose_print, info_print, error_print, QUIET_MODE
import asyncio
import aiohttp
import json
from read_chuck import DataLoader, TextProcessor, EvaluationConfig
from text_similarity import calculate_text_similarity
import logging
import openai
from dotenv import load_dotenv
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

# 导入 Ragas 相关模块
from ragas import evaluate, SingleTurnSample
from ragas.metrics import (
    Faithfulness,
    AnswerRelevancy,
    ContextPrecision,
    ContextRecall,
    ContextRelevance
)
from ragas import EvaluationDataset
from ragas.llms import LangchainLLMWrapper
from langchain_openai import ChatOpenAI
from ragas.testset import TestsetGenerator

# 加载环境变量
load_dotenv()

# 配置日志
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class StandardDatasetBuilder:
    """标准数据集构建器"""
    
    def __init__(self, knowledge_doc_dir: str = "knowledgeDoc", 
                 standard_dataset_path: str = "standardDataset/standardDataset_build.xlsx"):
        self.knowledge_doc_dir = Path(knowledge_doc_dir)
        self.standard_dataset_path = Path(standard_dataset_path)
        
        # 配置OpenAI客户端
        self.openai_client = openai.OpenAI(
            api_key=os.getenv('OPENAI_API_KEY') or os.getenv('QWEN_API_KEY'),
            base_url=os.getenv('OPENAI_API_BASE', 'https://api.openai.com/v1')
        )
        
        # 配置 Langchain LLM（使用稳定的采样参数）
        # 注意：Qwen API 要求参数显式指定，不能使用 model_kwargs
        self.llm = ChatOpenAI(
            model=os.getenv('MODEL_NAME', 'qwen-plus'),
            api_key=os.getenv('OPENAI_API_KEY') or os.getenv('QWEN_API_KEY'),
            base_url=os.getenv('OPENAI_API_BASE', 'https://api.openai.com/v1'),
            temperature=0.0,  # 降低到 0.0 以获得更稳定的输出
            top_p=0.1,  # 只从最高概率的 10% token 中选择
            max_tokens=2000  # 最大生成长度
        )
        
        # 配置 Ragas LLM（使用稳定的采样参数）
        # 注意：Qwen API 要求参数显式指定，不能使用 model_kwargs
        self.ragas_llm = LangchainLLMWrapper(
            ChatOpenAI(
                model=os.getenv('MODEL_NAME', 'qwen-plus'),
                api_key=os.getenv('OPENAI_API_KEY') or os.getenv('QWEN_API_KEY'),
                base_url=os.getenv('OPENAI_API_BASE', 'https://api.openai.com/v1'),
                temperature=0.0,  # 降低到 0.0 以获得更稳定的输出
                top_p=0.1,  # 只从最高概率的 10% token 中选择
                max_tokens=2000  # 最大生成长度
            )
        )
        
    def load_knowledge_documents(self) -> List[Dict[str, Any]]:
        """
        加载知识库文档
        
        Returns:
            List[Dict]: 文档列表，每个文档包含路径、内容、分块等信息
        """
        if not self.knowledge_doc_dir.exists():
            logger.warning(f"知识库目录不存在: {self.knowledge_doc_dir}")
            return []
        
        documents = []
        # 创建配置对象
        config = EvaluationConfig(
            api_key=os.getenv('OPENAI_API_KEY', ''),
            api_base=os.getenv('OPENAI_API_BASE', 'https://api.openai.com/v1'),
            model_name=os.getenv('MODEL_NAME', 'qwen-plus')
        )
        text_processor = TextProcessor(config)
        
        for file_path in self.knowledge_doc_dir.iterdir():
            if file_path.is_file():
                try:
                    # 读取文档内容
                    with open(file_path, 'r', encoding='utf-8') as f:
                        doc_content = f.read()
                    
                    if doc_content:
                        # 分块处理（保留原始文档内容，不清理标题）
                        chunks = text_processor.split_text_into_chunks(doc_content)
                        documents.append({
                            'path': str(file_path),
                            'name': file_path.name,
                            'content': doc_content,
                            'chunks': chunks
                        })
                        logger.info(f"加载文档: {file_path.name}, 分块数: {len(chunks)}")
                except Exception as e:
                    logger.error(f"加载文档失败 {file_path.name}: {e}")
        
        return documents
    
    
    def load_standard_dataset(self) -> Optional[pd.DataFrame]:
        """
        加载标准数据集
        
        Returns:
            pd.DataFrame: 标准数据集DataFrame
        """
        if not self.standard_dataset_path.exists():
            logger.error(f"标准数据集文件不存在: {self.standard_dataset_path}")
            return None
        
        try:
            df = pd.read_excel(self.standard_dataset_path)
            
            # 确保reference_contexts和reference列的数据类型为字符串
            if 'reference_contexts' in df.columns:
                df['reference_contexts'] = df['reference_contexts'].astype(str)
            if 'reference' in df.columns:
                df['reference'] = df['reference'].astype(str)
            
            logger.info(f"加载标准数据集: {len(df)} 行数据")
            return df
        except Exception as e:
            logger.error(f"加载标准数据集失败: {e}")
            return None
    
    async def generate_reference_answer(self, query: str, contexts: List[str]) -> Dict[str, Any]:
        """
        使用LLM生成标准答案
        
        Args:
            query: 用户查询
            contexts: 相关上下文
            
        Returns:
            Dict: 生成结果
        """
        try:
            # 构建提示词
            context_text = "\n\n".join(contexts)
            
            prompt = f"""基于以下上下文信息，为用户的查询生成标准答案。

用户查询: {query}

上下文信息:
{context_text}

请按照以下要求生成答案:
1. 基于提供的上下文信息回答用户查询
2. 答案要准确、完整、有逻辑性
3. 如果上下文中没有相关信息，请明确说明
4. 答案要简洁明了，避免冗余

重要格式要求:
- 直接给出答案内容，不要添加任何前缀
- 不要使用"标准答案："、"答案："、"回答："等前缀
- 直接开始回答，第一句话就是答案内容

请直接给出答案:"""

            # 调用LLM生成答案
            response = self.openai_client.chat.completions.create(
                model=os.getenv('MODEL_NAME', 'qwen-plus'),
                messages=[
                    {"role": "system", "content": "你是一个专业的问答助手，能够基于提供的上下文信息生成准确的标准答案。请直接给出答案内容，不要添加任何前缀如'标准答案：'、'答案：'等。"},
                    {"role": "user", "content": prompt}
                ],
                temperature=0.1,
                max_tokens=1000
            )
            
            reference_answer = response.choices[0].message.content.strip()
            
            # 清理可能包含的前缀字符串
            reference_answer = self.clean_reference_answer(reference_answer)
            
            # 打印生成的标准答案
            info_print(f"\n{'='*80}")
            info_print(f"📝 查询: {query}")
            info_print(f"{'='*80}")
            info_print(f"🤖 生成的标准答案:")
            info_print(f"{reference_answer}")
            info_print(f"{'='*80}\n")
            
            # 使用 LLM 选择最相关的上下文作为reference_contexts
            info_print(f"\n🔍 使用 LLM 选择相关上下文...")
            info_print(f"📊 传入的 contexts 数量: {len(contexts) if contexts else 0}")
            if contexts:
                info_print(f"📄 第一个分块示例: {contexts[0][:100]}...")
            relevant_contexts = await self.select_relevant_contexts_with_llm(query, contexts, max_contexts=10)
            
            # 打印所有构建出的分块
            info_print(f"\n{'='*80}")
            info_print(f"📚 reference_contexts构建出的所有分块:")
            info_print(f"{'='*80}")
            info_print(f"🔍 查询: {query}")
            info_print(f"📊 总上下文分块数: {len(contexts)}")
            info_print(f"🎯 选择的相关分块数: {len(relevant_contexts)}")
            info_print(f"{'='*80}")
            
            for i, chunk in enumerate(relevant_contexts):
                info_print(f"\n📄 分块 {i+1}:")
                info_print(f"{'-'*60}")
                info_print(f"{chunk}")
                info_print(f"{'-'*60}")
            
            info_print(f"\n{'='*80}")
            info_print(f"📋 分块构建完成")
            info_print(f"{'='*80}\n")
            
            # 使用 Ragas 评估生成的标准答案质量
            info_print(f"\n🔍 使用 Ragas 评估标准答案质量...")
            ragas_scores = await self.evaluate_with_ragas_metrics(query, reference_answer, relevant_contexts)
            
            return {
                "reference": reference_answer,
                "reference_contexts": relevant_contexts,
                "ragas_scores": ragas_scores
            }
            
        except Exception as e:
            logger.error(f"生成标准答案失败: {e}")
            return None
    
    def calculate_relevance_score(self, query: str, chunk: str) -> float:
        """
        计算查询与分块的相关性得分
        使用app.py中的calculate_text_similarity函数
        
        Args:
            query: 用户查询
            chunk: 分块内容
            
        Returns:
            float: 相关性得分 (0-1)
        """
        if not query or not chunk:
            return 0.0
        
        # 使用app.py中的相似度计算函数
        similarity = calculate_text_similarity(query, chunk)
        
        return similarity
    
    async def is_chunk_relevant_with_llm(self, query: str, chunk: str) -> bool:
        """
        使用 LLM 判断分块是否与查询相关
        
        Args:
            query: 用户查询
            chunk: 文档分块
            
        Returns:
            bool: 是否相关
        """
        try:
            prompt = f"""
用户问题: {query}

文档分块: {chunk}

# 角色
自然语言语义相关性判定专家，擅长运用先进的自然语言处理技术，精准判断文本块与给定用户输入之间的语义相关性。

# 目标
1. 判定当前分块是否与“用户问题”语义相关。
2. 确定该分块能否作为回答“用户问题”的内容素材。

# 技能
1. 熟练掌握自然语言处理中的语义分析技术。
2. 具备文本相似度计算的能力。

# 工作流程
1. 仔细理解“用户问题”的语义、核心意图、关键词。
2. 对当前分块的内容进行深入剖析，提取关键信息、关键词。
3. 运用语义分析和相似度计算方法，对比分块与“用户问题”的语义、关键词。
4. 根据对比结果，判断分块是否与“用户问题”语义、关键词语义相关。
5. 确定该分块是否可作为回答“用户问题”的内容素材。

# 约束
1. 必须严格依据语义相关性进行判断，不得受其他无关因素干扰。
2. 禁止主观臆断，判断结果需有合理的分析依据。

# 输出格式
只回答“相关”或“不相关”，不要添加其他内容。
"""
            
            response = await self.llm.ainvoke(prompt)
            result = response.content.strip()
            result_lower = result.lower()
            
            info_print(f"🤖 LLM 响应: {result}")
            
            # 判断是否相关
            relevant_keywords = ['相关', 'relevant', 'yes', '是', '可以', '能']
            irrelevant_keywords = ['不相关', 'irrelevant', 'no', '不是', '不能', '不可以']
            
            # 优先检查不相关的关键词
            if any(keyword in result_lower for keyword in irrelevant_keywords):
                info_print(f"❌ 判定为不相关")
                return False
            
            # 然后检查相关的关键词
            if any(keyword in result_lower for keyword in relevant_keywords):
                info_print(f"✅ 判定为相关")
                return True
            
            # 如果都没有匹配到，默认认为不相关
            info_print(f"⚠️ LLM 响应不明确: {result}")
            return False
            
        except Exception as e:
            logger.error(f"LLM 判断分块相关性失败: {e}")
            info_print(f"⚠️ LLM 调用失败: {e}")
            # LLM 调用失败时，默认认为不相关，避免选择错误的分块
            info_print(f"🎯 默认判断结果: 不相关")
            return False
    
    async def select_relevant_contexts_with_llm(self, query: str, contexts: List[str], max_contexts: int = 10) -> List[str]:
        """
        使用 LLM 选择与查询最相关的上下文分块
        
        Args:
            query: 用户查询
            contexts: 所有分块列表
            max_contexts: 最大选择的分块数量
            
        Returns:
            List[str]: 最相关的分块列表
        """
        if not contexts:
            return []
        
        info_print(f"\n🔍 使用 LLM 判断分块相关性...")
        info_print(f"📊 查询: {query}")
        info_print(f"📚 总分块数: {len(contexts)}")
        info_print(f"🎯 最大选择数: {max_contexts}")
        
        relevant_contexts = []
        
        for i, chunk in enumerate(contexts):
            info_print(f"🔍 正在判断分块 {i+1}/{len(contexts)}...")
            info_print(f"📄 分块内容: {chunk[:100]}...")
            
            # 使用 LLM 判断相关性
            is_relevant = await self.is_chunk_relevant_with_llm(query, chunk)
            
            if is_relevant:
                relevant_contexts.append(chunk)
                info_print(f"✅ 分块 {i+1} 被判定为相关")
                
                # 如果已经达到最大数量，停止处理
                if len(relevant_contexts) >= max_contexts:
                    info_print(f"🎯 已达到最大选择数量 {max_contexts}")
                    break
            else:
                info_print(f"❌ 分块 {i+1} 被判定为不相关")
        
        info_print(f"\n🎯 最终选择了 {len(relevant_contexts)} 个相关分块")
        return relevant_contexts
    
    async def generate_testset_with_ragas(self, knowledge_docs: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """
        使用 Ragas 的 TestsetGenerator 生成测试数据集
        
        Args:
            knowledge_docs: 知识库文档列表
            
        Returns:
            List[Dict]: 生成的测试数据
        """
        try:
            # 合并所有文档内容
            all_content = []
            for doc in knowledge_docs:
                all_content.extend(doc['chunks'])
            
            # 创建 TestsetGenerator
            generator = TestsetGenerator.with_openai(
                generator_llm=self.ragas_llm,
                critic_llm=self.ragas_llm
            )
            
            # 生成测试集
            testset = await generator.agenerate(
                documents=all_content,
                test_size=10,  # 生成10个测试样本
                with_deep_eval=True,
                raise_exceptions=False
            )
            
            logger.info(f"使用 Ragas 生成了 {len(testset)} 个测试样本")
            return testset
            
        except Exception as e:
            logger.error(f"使用 Ragas 生成测试集失败: {e}")
            return []
    
    async def evaluate_with_ragas_metrics(self, query: str, answer: str, contexts: List[str]) -> Dict[str, float]:
        """
        简化的评估方法，不进行复杂的评估计算
        
        Args:
            query: 用户查询
            answer: 生成的标准答案
            contexts: 相关上下文
            
        Returns:
            Dict[str, float]: 空的评估指标（不进行评估）
        """
        info_print(f"\n🔍 跳过复杂评估，直接返回空结果...")
        return {}

    def clean_reference_answer(self, answer: str) -> str:
        """
        清理标准答案中的前缀字符串
        
        Args:
            answer: 原始答案
            
        Returns:
            str: 清理后的答案
        """
        if not answer:
            return answer
        
        # 需要清理的前缀字符串列表
        prefixes_to_remove = [
            "标准答案：",
            "标准答案:",
            "答案：",
            "答案:",
            "回答：",
            "回答:",
            "Reference Answer:",
            "Reference Answer：",
            "Answer:",
            "Answer："
        ]
        
        # 清理前缀
        cleaned_answer = answer.strip()
        for prefix in prefixes_to_remove:
            if cleaned_answer.startswith(prefix):
                cleaned_answer = cleaned_answer[len(prefix):].strip()
                break
        
        return cleaned_answer
    
    
    
    
    
    def format_contexts(self, contexts: List[str]) -> str:
        """
        格式化上下文，分块之间用空行分隔
        
        Args:
            contexts: 上下文列表
            
        Returns:
            str: 格式化后的上下文
        """
        info_print(f"🔧 format_contexts 调试:")
        info_print(f"  输入 contexts 类型: {type(contexts)}")
        info_print(f"  输入 contexts 长度: {len(contexts) if contexts else 0}")
        
        if not contexts:
            info_print(f"  ⚠️ contexts 为空，返回空字符串")
            return ""
        
        # 过滤掉空的分块
        filtered_contexts = [ctx for ctx in contexts if ctx.strip()]
        info_print(f"  过滤后分块数量: {len(filtered_contexts)}")
        
        if filtered_contexts:
            info_print(f"  第一个分块示例: {filtered_contexts[0][:50]}...")
        
        # 使用空行连接分块
        result = "\n\n".join(filtered_contexts)
        info_print(f"  格式化结果长度: {len(result)}")
        return result
    
    async def build_reference_data(self, df: pd.DataFrame, 
                                 knowledge_docs: List[Dict[str, Any]]) -> pd.DataFrame:
        """
        构建标准答案数据
        
        Args:
            df: 标准数据集DataFrame
            knowledge_docs: 知识库文档列表
            
        Returns:
            pd.DataFrame: 更新后的DataFrame
        """
        # 合并所有知识库文档的分块
        all_chunks = []
        for doc in knowledge_docs:
            all_chunks.extend(doc['chunks'])
        
        logger.info(f"总共有 {len(all_chunks)} 个知识库分块")
        info_print(f"📊 知识库分块统计:")
        info_print(f"  📚 文档数量: {len(knowledge_docs)}")
        info_print(f"  📄 总分块数: {len(all_chunks)}")
        if all_chunks:
            info_print(f"  📝 第一个分块示例: {all_chunks[0][:100]}...")
        else:
            info_print(f"  ⚠️ 没有找到任何分块！")
        
        for index, row in df.iterrows():
            try:
                query = row['user_input']
                logger.info(f"处理查询 {index + 1}/{len(df)}: {query[:50]}...")
                
                # 生成标准答案
                result = await self.generate_reference_answer(query, all_chunks)
                
                if result:
                    # 更新reference_contexts和reference
                    reference_contexts = result.get('reference_contexts', [])
                    info_print(f"🔧 准备保存 reference_contexts:")
                    info_print(f"  reference_contexts 类型: {type(reference_contexts)}")
                    info_print(f"  reference_contexts 长度: {len(reference_contexts) if reference_contexts else 0}")
                    
                    formatted_contexts = self.format_contexts(reference_contexts)
                    df.at[index, 'reference_contexts'] = str(formatted_contexts)
                    df.at[index, 'reference'] = str(result.get('reference', ''))
                    
                    info_print(f"💾 保存到 DataFrame:")
                    info_print(f"  reference_contexts 列值长度: {len(str(formatted_contexts))}")
                    info_print(f"  reference 列值长度: {len(str(result.get('reference', '')))}")
                    
                    logger.info(f"✅ 成功生成标准答案: {index + 1}/{len(df)}")
                    info_print(f"✅ 第 {index + 1}/{len(df)} 条标准答案生成完成")
                else:
                    logger.warning(f"❌ 生成标准答案失败: {index + 1}")
                    info_print(f"❌ 第 {index + 1}/{len(df)} 条标准答案生成失败")
                    
            except Exception as e:
                logger.error(f"处理查询失败 {index + 1}: {e}")
                continue
        
        return df
    
    def save_updated_dataset(self, df: pd.DataFrame) -> bool:
        """
        保存更新后的数据集
        
        Args:
            df: 更新后的DataFrame
            
        Returns:
            bool: 保存是否成功
        """
        try:
            # 确保目录存在
            self.standard_dataset_path.parent.mkdir(parents=True, exist_ok=True)
            
            # 确保reference_contexts和reference列的数据类型为字符串
            if 'reference_contexts' in df.columns:
                df['reference_contexts'] = df['reference_contexts'].astype(str)
            if 'reference' in df.columns:
                df['reference'] = df['reference'].astype(str)
            
            # 如果原文件被占用，使用备份文件名
            save_path = self.standard_dataset_path
            if save_path.exists():
                backup_path = save_path.parent / f"{save_path.stem}_updated{save_path.suffix}"
                save_path = backup_path
            
            # 保存文件
            df.to_excel(save_path, index=False)
            
            # 使用 openpyxl 调整行高和列宽
            self._adjust_excel_formatting(save_path)
            
            logger.info(f"数据集保存成功: {save_path}")
            return True
            
        except Exception as e:
            logger.error(f"保存数据集失败: {e}")
            return False
    
    def _adjust_excel_formatting(self, file_path: Path) -> None:
        """
        调整 Excel 文件的格式，让单元格跟随内容自适应显示
        
        Args:
            file_path: Excel 文件路径
        """
        try:
            # 加载工作簿
            workbook = load_workbook(file_path)
            worksheet = workbook.active
            
            # 1. 设置单元格格式（自动换行和对齐）
            for row in worksheet.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(
                        wrap_text=True,      # 自动换行
                        vertical='top',      # 垂直顶部对齐
                        horizontal='left'    # 水平左对齐
                    )
            
            # 2. 自适应列宽
            for col in worksheet.columns:
                # 获取列索引（从1开始）
                col_idx = col[0].column
                # 计算该列中最长内容的长度
                max_length = max(len(str(cell.value)) for cell in col if cell.value)
                # 设置列宽（加2是为了留一些余量，限制最大宽度为50）
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[get_column_letter(col_idx)].width = adjusted_width
            
            # 3. 自适应行高
            for row_idx, row in enumerate(worksheet.iter_rows(), 1):
                max_height = 15  # 默认行高
                
                for cell in row:
                    if cell.value:
                        text = str(cell.value)
                        # 计算文本行数（基于换行符）
                        lines = text.count('\n') + 1
                        
                        # 根据列宽估算可能的换行数
                        col_letter = get_column_letter(cell.column)
                        col_width = worksheet.column_dimensions[col_letter].width
                        if col_width > 0:
                            # 估算每行能容纳的字符数（中文字符按2个字符计算）
                            chars_per_line = int(col_width * 1.5)
                            estimated_lines = len(text) // chars_per_line + 1
                            lines = max(lines, estimated_lines)
                        
                        # 每行大约15像素高度
                        cell_height = lines * 15
                        max_height = max(max_height, cell_height)
                
                # 设置行高（限制最大高度为200）
                adjusted_height = min(max(max_height, 15), 200)
                worksheet.row_dimensions[row_idx].height = adjusted_height
            
            # 保存文件
            workbook.save(file_path)
            logger.info(f"Excel 格式调整完成，单元格已自适应内容: {file_path}")
            
        except Exception as e:
            logger.error(f"调整 Excel 格式失败: {e}")
    
    async def build_dataset(self) -> Dict[str, Any]:
        """
        构建标准数据集的主函数
        
        Returns:
            Dict: 构建结果
        """
        try:
            logger.info("开始构建标准数据集...")
            
            # 1. 加载知识库文档
            knowledge_docs = self.load_knowledge_documents()
            if not knowledge_docs:
                return {
                    "success": False,
                    "message": "没有找到知识库文档"
                }
            
            # 2. 加载标准数据集
            df = self.load_standard_dataset()
            if df is None:
                return {
                    "success": False,
                    "message": "加载标准数据集失败"
                }
            
            # 3. 构建标准答案数据
            updated_df = await self.build_reference_data(df, knowledge_docs)
            
            # 4. 保存更新后的数据集
            if self.save_updated_dataset(updated_df):
                # 打印所有生成的标准答案摘要
                info_print(f"\n{'='*100}")
                info_print(f"📊 标准数据集构建完成摘要")
                info_print(f"{'='*100}")
                info_print(f"📈 总处理查询数: {len(df)}")
                info_print(f"📚 知识库文档数: {len(knowledge_docs)}")
                info_print(f"💾 保存路径: {self.standard_dataset_path}")
                info_print(f"{'='*100}")
                
                # 打印所有生成的标准答案
                info_print(f"\n📋 所有生成的标准答案:")
                info_print(f"{'='*100}")
                for index, row in updated_df.iterrows():
                    if pd.notna(row.get('reference', '')):
                        info_print(f"\n🔸 查询 {index + 1}: {row['user_input'][:50]}...")
                        
                        # 打印所有reference_contexts的分块
                        reference_contexts = row.get('reference_contexts', '')
                        if reference_contexts and reference_contexts != 'nan':
                            info_print(f"\n📚 reference_contexts 分块内容:")
                            info_print(f"{'='*80}")
                            
                            # 将reference_contexts按分块分割（假设用双换行符分隔）
                            if isinstance(reference_contexts, str):
                                chunks = reference_contexts.split('\n\n')
                                for i, chunk in enumerate(chunks):
                                    if chunk.strip():  # 只显示非空分块
                                        info_print(f"\n📄 分块 {i+1}:")
                                        info_print(f"{'-'*60}")
                                        info_print(f"{chunk.strip()}")
                                        info_print(f"{'-'*60}")
                            else:
                                info_print(f"reference_contexts 不是字符串格式: {type(reference_contexts)}")
                        else:
                            info_print(f"\n⚠️  reference_contexts 为空或无效")
                        
                        info_print(f"\n📝 标准答案: {row['reference']}")
                        info_print(f"{'-'*80}")
                
                return {
                    "success": True,
                    "message": f"数据集构建成功！处理了 {len(df)} 条查询",
                    "processed_count": len(df),
                    "knowledge_docs_count": len(knowledge_docs)
                }
            else:
                return {
                    "success": False,
                    "message": "保存数据集失败"
                }
                
        except Exception as e:
            logger.error(f"构建数据集异常: {e}")
            return {
                "success": False,
                "message": f"构建数据集失败: {str(e)}"
            }

# 便捷函数
async def build_standard_dataset() -> Dict[str, Any]:
    """
    构建标准数据集的便捷函数
    
    Returns:
        Dict: 构建结果
    """
    builder = StandardDatasetBuilder()
    return await builder.build_dataset()

# 测试函数
async def test_build_dataset():
    """测试构建数据集功能"""
    info_print("🔍 测试标准数据集构建功能...")
    
    result = await build_standard_dataset()
    info_print(f"📋 构建结果: {result}")
    
    if result["success"]:
        info_print("✅ 数据集构建成功")
    else:
        info_print("❌ 数据集构建失败")

if __name__ == "__main__":
    asyncio.run(test_build_dataset())
